import io
import os
import mimetypes
import sqlite3
import json
import uuid
import datetime
import random
import string
from flask import Flask, send_from_directory, request, send_file, session, redirect, url_for, jsonify
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from functools import wraps
from datetime import timedelta
from werkzeug.security import generate_password_hash, check_password_hash
from flask_principal import Principal, Permission, RoleNeed, Identity, identity_loaded, identity_changed

# Ensure JS/CSS mimetypes are registered correctly for ES modules
mimetypes.add_type('application/javascript', '.js')
mimetypes.add_type('text/css', '.css')

app = Flask(__name__, static_folder='.', static_url_path='')
principal = Principal(app)
app.secret_key = os.environ.get('SECRET_KEY', 'PROJECTCRM_SECURE_SECRET_2026_KEY')
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(days=30)
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'

# Configure Upload Folders
try:
    os.makedirs('/data/uploads/profiles', exist_ok=True)
    UPLOAD_BASE_DIR = '/data/uploads'
except (PermissionError, OSError):
    UPLOAD_BASE_DIR = os.path.join(os.path.dirname(__file__), 'static', 'uploads')
    os.makedirs(os.path.join(UPLOAD_BASE_DIR, 'profiles'), exist_ok=True)

app.config['UPLOAD_FOLDER'] = UPLOAD_BASE_DIR

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('user_id'):
            return {"error": "Yetkisiz erişim. Lütfen giriş yapın."}, 401
        return f(*args, **kwargs)
    return decorated_function

# Flask-Principal Identity Loader
@principal.identity_loader
def read_identity_from_session():
    if 'user_id' in session:
        return Identity(session['user_id'])
    return None

# Load Roles & Permissions into Identity
def on_identity_loaded(sender, identity):
    identity.user = identity.id
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute('SELECT role FROM users WHERE uid = ?', (identity.id,))
        user_row = cursor.fetchone()
        if user_row:
            role = user_row['role'] or 'agent'
            identity.provides.add(RoleNeed(role))
            
            # Load permissions from rol_yetkileri
            cursor.execute('SELECT * FROM rol_yetkileri WHERE role = ?', (role,))
            perm_row = cursor.fetchone()
            if perm_row:
                perm_dict = dict(perm_row)
                for perm_name, has_val in perm_dict.items():
                    if perm_name != 'role' and has_val:
                        identity.provides.add(RoleNeed(f"perm:{perm_name}"))
        conn.close()
    except Exception as e:
        print(f"Error loading identity permissions: {e}", flush=True)

identity_loaded.connect(on_identity_loaded, app)

# 403 Forbidden Page HTML Template
def render_forbidden_html():
    return """
    <!DOCTYPE html>
    <html lang="tr">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Erişim Engellendi (403) - PROJECTCRM</title>
        <link href="https://fonts.googleapis.com/css2?family=Outfit:wght@400;600;700&display=swap" rel="stylesheet">
        <style>
            :root {
                --bg-color: #0b0f19;
                --card-bg: rgba(30, 41, 59, 0.4);
                --border-color: rgba(255, 255, 255, 0.08);
                --primary-color: #6366f1;
                --text-primary: #f8fafc;
                --text-secondary: #94a3b8;
            }
            body {
                background-color: var(--bg-color);
                color: var(--text-primary);
                font-family: 'Outfit', sans-serif;
                display: flex;
                align-items: center;
                justify-content: center;
                height: 100vh;
                margin: 0;
                overflow: hidden;
            }
            .container {
                text-align: center;
                background: var(--card-bg);
                border: 1px solid var(--border-color);
                padding: 40px;
                border-radius: 16px;
                backdrop-filter: blur(12px);
                max-width: 450px;
                width: 90%;
                box-shadow: 0 8px 32px 0 rgba(0, 0, 0, 0.37);
            }
            .icon {
                font-size: 64px;
                margin-bottom: 24px;
                animation: pulse 2s infinite ease-in-out;
            }
            h1 {
                font-size: 24px;
                font-weight: 700;
                margin: 0 0 12px 0;
            }
            p {
                font-size: 14px;
                color: var(--text-secondary);
                margin: 0 0 24px 0;
                line-height: 1.6;
            }
            .btn {
                background: var(--primary-color);
                color: white;
                border: none;
                padding: 12px 24px;
                font-size: 14px;
                font-weight: 600;
                border-radius: 8px;
                cursor: pointer;
                text-decoration: none;
                transition: all 0.2s ease;
                display: inline-block;
            }
            .btn:hover {
                opacity: 0.9;
                transform: translateY(-1px);
            }
            @keyframes pulse {
                0% { transform: scale(1); }
                50% { transform: scale(1.05); }
                100% { transform: scale(1); }
            }
        </style>
    </head>
    <body>
        <div class="container">
            <div class="icon">🔒</div>
            <h1>Erişim Yetkiniz Bulunmamaktadır</h1>
            <p>Bu sayfaya veya işleme erişmek için gerekli yetki izinlerine sahip değilsiniz. Lütfen acente yöneticiniz ile iletişime geçin.</p>
            <a href="/" class="btn">Ana Sayfaya Dön</a>
        </div>
    </body>
    </html>
    """, 403

# Dynamic Roles Accepted Decorator
def roles_accepted(*requirements):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not session.get('user_id'):
                return {"error": "Yetkisiz erişim. Lütfen giriş yapın."}, 401
                
            has_access = False
            try:
                conn = get_db()
                cursor = conn.cursor()
                cursor.execute('SELECT role FROM users WHERE uid = ?', (session.get('user_id'),))
                user_row = cursor.fetchone()
                if user_row:
                    role = user_row['role'] or 'agent'
                    if role in requirements:
                        has_access = True
                    else:
                        cursor.execute('SELECT * FROM rol_yetkileri WHERE role = ?', (role,))
                        perm_row = cursor.fetchone()
                        if perm_row:
                            perm_dict = dict(perm_row)
                            for req in requirements:
                                if req in perm_dict and perm_dict[req]:
                                    has_access = True
                                    break
                conn.close()
            except Exception as e:
                print(f"Error checking roles_accepted: {e}", flush=True)
                
            if not has_access:
                if request.accept_mimetypes.accept_html and not request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return render_forbidden_html()
                else:
                    return {"error": "Bu işlem için yetkiniz bulunmamaktadır."}, 403
            return f(*args, **kwargs)
        return decorated_function
    return decorator

# Dynamic Roles Required Decorator
def roles_required(*requirements):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not session.get('user_id'):
                return {"error": "Yetkisiz erişim. Lütfen giriş yapın."}, 401
                
            has_access = True
            try:
                conn = get_db()
                cursor = conn.cursor()
                cursor.execute('SELECT role FROM users WHERE uid = ?', (session.get('user_id'),))
                user_row = cursor.fetchone()
                if user_row:
                    role = user_row['role'] or 'agent'
                    cursor.execute('SELECT * FROM rol_yetkileri WHERE role = ?', (role,))
                    perm_row = cursor.fetchone()
                    perm_dict = dict(perm_row) if perm_row else {}
                    
                    for req in requirements:
                        if req in ['admin', 'agent', 'assistant']:
                            if role != req:
                                has_access = False
                                break
                        else:
                            if not perm_dict.get(req):
                                has_access = False
                                break
                else:
                    has_access = False
                conn.close()
            except Exception as e:
                print(f"Error checking roles_required: {e}", flush=True)
                has_access = False
                
            if not has_access:
                if request.accept_mimetypes.accept_html and not request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return render_forbidden_html()
                else:
                    return {"error": "Bu işlem için yetkiniz bulunmamaktadır."}, 403
            return f(*args, **kwargs)
        return decorated_function
    return decorator


@app.route('/')
def index():
    return send_from_directory('.', 'index.html')

@app.route('/login')
def login_route():
    if session.get('user_id'):
        return redirect('/')
    return redirect('/#auth')

@app.route('/logout')
def logout_route():
    session.clear()
    response = redirect('/login')
    response.set_cookie(app.config.get('SESSION_COOKIE_NAME', 'session'), '', expires=0)
    return response

@app.route('/api/auth/logout', methods=['POST', 'GET'])
def api_auth_logout():
    session.clear()
    response = jsonify({"success": True})
    response.set_cookie(app.config.get('SESSION_COOKIE_NAME', 'session'), '', expires=0)
    return response

@app.route('/uploads/<path:filename>')
def serve_uploads(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)

@app.route('/<path:path>')
def serve_static(path):
    return send_from_directory('.', path)

@app.after_request
def add_header(response):
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    return response

@app.route('/api/export', methods=['POST'])
@login_required
def export_excel():
    try:
        data = request.get_json() or {}
        customers = data.get('customers', [])
        portfolios = data.get('portfolios', [])

        # Filter buyers and sellers
        buyers = [c for c in customers if c.get('type') == 'Alıcı']
        sellers = [c for c in customers if c.get('type') == 'Satıcı']

        # Create Workbook
        wb = openpyxl.Workbook()
        # Remove default sheet
        default_sheet = wb.active
        wb.remove(default_sheet)

        # Styles
        font_family = "Segoe UI"
        title_font = Font(name=font_family, size=14, bold=True, color="1E293B")
        section_font = Font(name=font_family, size=12, bold=True, color="0F172A")
        header_font = Font(name=font_family, size=10, bold=True, color="FFFFFF")
        data_font = Font(name=font_family, size=10, color="334155")
        
        # Dark Slate for portfolios and Medium Slate for seller titles
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_buyer_fill = PatternFill(start_color="047857", end_color="047857", fill_type="solid") # Dark Emerald
        header_seller_fill = PatternFill(start_color="475569", end_color="475569", fill_type="solid") # Medium Slate
        
        thin_border_side = Side(style='thin', color='CBD5E1')
        thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

        # Alignments
        align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
        align_right = Alignment(horizontal='right', vertical='center', wrap_text=True)

        # -------------------------------------------------------------
        # 1. Alıcı Takip Listesi Sheet
        # -------------------------------------------------------------
        ws1 = wb.create_sheet(title="Alıcı Takip Listesi")
        ws1.views.sheetView[0].showGridLines = True
        
        # Title Block
        ws1.cell(row=1, column=1, value="ALICI MÜŞTERİ TAKİP LİSTESİ").font = title_font
        ws1.row_dimensions[1].height = 30
        
        headers_buyer = [
            "Müşteri Adı", "Telefon", "E-posta", "Doğum Tarihi", 
            "Bütçe", "Hedef Bölge", "Aranan Oda", "Konut Tipi", 
            "Finansman Tipi", "Satın Alma Amacı", "Yabancı Satış", "Aciliyet Durumu",
            "Danışman", "Notlar"
        ]
        
        for col_num, header in enumerate(headers_buyer, 1):
            cell = ws1.cell(row=3, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_buyer_fill
            cell.alignment = align_center
            cell.border = thin_border
        
        ws1.row_dimensions[3].height = 26
        
        row_idx = 4
        for b in buyers:
            ws1.cell(row=row_idx, column=1, value=b.get('name', '')).alignment = align_left
            ws1.cell(row=row_idx, column=2, value=b.get('phone', '')).alignment = align_center
            ws1.cell(row=row_idx, column=3, value=b.get('email', '')).alignment = align_left
            ws1.cell(row=row_idx, column=4, value=b.get('birthDate', '')).alignment = align_center
            
            # Budget
            budget_val = b.get('budget', 0)
            try:
                budget_val = float(budget_val)
            except (ValueError, TypeError):
                budget_val = 0
            budget_cell = ws1.cell(row=row_idx, column=5, value=budget_val)
            budget_cell.alignment = align_right
            budget_cell.number_format = '₺#,##0'
            
            ws1.cell(row=row_idx, column=6, value=b.get('searchLocation', '')).alignment = align_left
            ws1.cell(row=row_idx, column=7, value=b.get('searchRooms', '')).alignment = align_center
            ws1.cell(row=row_idx, column=8, value=b.get('searchPropertyType', '')).alignment = align_left
            ws1.cell(row=row_idx, column=9, value=b.get('finansman_tipi', '')).alignment = align_center
            ws1.cell(row=row_idx, column=10, value=b.get('satin_alma_amaci', '')).alignment = align_center
            ws1.cell(row=row_idx, column=11, value=b.get('yabanci_satis', '')).alignment = align_center
            ws1.cell(row=row_idx, column=12, value=b.get('aciliyet_durumu', '')).alignment = align_center
            ws1.cell(row=row_idx, column=13, value=b.get('createdByName', '')).alignment = align_left
            ws1.cell(row=row_idx, column=14, value=b.get('notes', '')).alignment = align_left
            
            for col_num in range(1, len(headers_buyer) + 1):
                c_cell = ws1.cell(row=row_idx, column=col_num)
                c_cell.font = data_font
                c_cell.border = thin_border
                
            ws1.row_dimensions[row_idx].height = 20
            row_idx += 1
            
        # Auto-fit Column Widths for WS1
        for col in ws1.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            # Skip title row in calculating widths
            for cell in col[2:]: # from headers downwards
                if cell.value:
                    if cell.number_format == '₺#,##0':
                        val_str = f"₺{int(cell.value):,}"
                    else:
                        val_str = str(cell.value)
                    max_len = max(max_len, len(val_str))
            ws1.column_dimensions[col_letter].width = max(max_len + 4, 12)

        # -------------------------------------------------------------
        # 2. Satıcı & Portföy Listesi Sheet
        # -------------------------------------------------------------
        ws2 = wb.create_sheet(title="Satıcı & Portföy Listesi")
        ws2.views.sheetView[0].showGridLines = True
        
        # Title Table 1
        ws2.cell(row=1, column=1, value="PORTFÖY LİSTESİ").font = section_font
        ws2.row_dimensions[1].height = 24
        
        headers_portfolio = [
            "Portföy Başlığı", "Fiyat", "İşlem Tipi", "Konut Tipi", "Oda Sayısı", 
            "Alan (m²)", "İlçe", "Mahalle", "Isınma", "Bina Yaşı", "Durum",
            "Sözleşme Tipi", "Sözleşme Bitiş Tarihi", "Mülk Durumu", "Tapu Notları", "Danışman"
        ]
        
        for col_num, header in enumerate(headers_portfolio, 1):
            cell = ws2.cell(row=3, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center
            cell.border = thin_border
            
        ws2.row_dimensions[3].height = 26
        
        row_idx = 4
        for p in portfolios:
            ws2.cell(row=row_idx, column=1, value=p.get('title', '')).alignment = align_left
            
            # Price
            price_val = p.get('price', 0)
            try:
                price_val = float(price_val)
            except (ValueError, TypeError):
                price_val = 0
            price_cell = ws2.cell(row=row_idx, column=2, value=price_val)
            price_cell.alignment = align_right
            price_cell.number_format = '₺#,##0'
            
            ws2.cell(row=row_idx, column=3, value=p.get('type', '')).alignment = align_center
            ws2.cell(row=row_idx, column=4, value=p.get('propertyType', '')).alignment = align_left
            ws2.cell(row=row_idx, column=5, value=p.get('rooms', '')).alignment = align_center
            
            # Area
            area_val = p.get('area', '')
            try:
                area_val = int(area_val)
            except (ValueError, TypeError):
                pass
            ws2.cell(row=row_idx, column=6, value=area_val).alignment = align_center
            
            ws2.cell(row=row_idx, column=7, value=p.get('district', '')).alignment = align_left
            ws2.cell(row=row_idx, column=8, value=p.get('neighborhood', '')).alignment = align_left
            ws2.cell(row=row_idx, column=9, value=p.get('heating', '')).alignment = align_left
            ws2.cell(row=row_idx, column=10, value=p.get('age', '')).alignment = align_center
            ws2.cell(row=row_idx, column=11, value=p.get('status', '')).alignment = align_center
            ws2.cell(row=row_idx, column=12, value=p.get('sozlesme_tipi', '')).alignment = align_center
            ws2.cell(row=row_idx, column=13, value=p.get('sozlesme_bitis_tarihi', '')).alignment = align_center
            ws2.cell(row=row_idx, column=14, value=p.get('mulk_durumu', '')).alignment = align_center
            ws2.cell(row=row_idx, column=15, value=p.get('tapu_durumu_notlari', '')).alignment = align_left
            ws2.cell(row=row_idx, column=16, value=p.get('createdByName', '')).alignment = align_left
            
            for col_num in range(1, len(headers_portfolio) + 1):
                c_cell = ws2.cell(row=row_idx, column=col_num)
                c_cell.font = data_font
                c_cell.border = thin_border
                
            ws2.row_dimensions[row_idx].height = 20
            row_idx += 1
            
        # Spacer row
        row_idx += 2
        
        # Title Table 2
        ws2.cell(row=row_idx, column=1, value="SATICI MÜŞTERİ LİSTESİ").font = section_font
        ws2.row_dimensions[row_idx].height = 24
        row_idx += 2
        
        headers_seller = [
            "Müşteri Adı", "Telefon", "E-posta", "Doğum Tarihi", 
            "Sözleşme Tipi", "Sözleşme Bitiş Tarihi", "Mülk Durumu", "Tapu Notları", 
            "Danışman", "Notlar"
        ]
        
        header_seller_row = row_idx
        for col_num, header in enumerate(headers_seller, 1):
            cell = ws2.cell(row=header_seller_row, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_seller_fill
            cell.alignment = align_center
            cell.border = thin_border
            
        ws2.row_dimensions[header_seller_row].height = 26
        
        row_idx += 1
        for s in sellers:
            ws2.cell(row=row_idx, column=1, value=s.get('name', '')).alignment = align_left
            ws2.cell(row=row_idx, column=2, value=s.get('phone', '')).alignment = align_center
            ws2.cell(row=row_idx, column=3, value=s.get('email', '')).alignment = align_left
            ws2.cell(row=row_idx, column=4, value=s.get('birthDate', '')).alignment = align_center
            ws2.cell(row=row_idx, column=5, value=s.get('sozlesme_tipi', '')).alignment = align_center
            ws2.cell(row=row_idx, column=6, value=s.get('sozlesme_bitis_tarihi', '')).alignment = align_center
            ws2.cell(row=row_idx, column=7, value=s.get('mulk_durumu', '')).alignment = align_center
            ws2.cell(row=row_idx, column=8, value=s.get('tapu_durumu_notlari', '')).alignment = align_left
            ws2.cell(row=row_idx, column=9, value=s.get('createdByName', '')).alignment = align_left
            ws2.cell(row=row_idx, column=10, value=s.get('notes', '')).alignment = align_left
            
            for col_num in range(1, len(headers_seller) + 1):
                c_cell = ws2.cell(row=row_idx, column=col_num)
                c_cell.font = data_font
                c_cell.border = thin_border
                
            ws2.row_dimensions[row_idx].height = 20
            row_idx += 1
            
        # Auto-fit Column Widths for WS2 (calculated across both tables)
        max_cols = max(len(headers_portfolio), len(headers_seller))
        for col_num in range(1, max_cols + 1):
            col_letter = get_column_letter(col_num)
            max_len = 0
            for cell in ws2[col_letter]:
                # Skip title rows, spacer rows and header rows for width calculations
                if cell.row in [1, 2, header_seller_row - 2, header_seller_row - 1]:
                    continue
                if cell.value is not None:
                    if cell.number_format == '₺#,##0':
                        val_str = f"₺{int(cell.value):,}"
                    else:
                        val_str = str(cell.value)
                    max_len = max(max_len, len(val_str))
            ws2.column_dimensions[col_letter].width = max(max_len + 4, 12)

        # Save to BytesIO
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        
        return send_file(
            out,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="crm_verileri_export.xlsx"
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        return {"error": str(e)}, 500

DATABASE_PATH = os.environ.get('DATABASE_PATH', os.path.join(os.path.dirname(__file__), 'projectcrm.db'))
# Ensure the parent directory of the database file exists
db_dir = os.path.dirname(DATABASE_PATH)
if db_dir and not os.path.exists(db_dir):
    os.makedirs(db_dir, exist_ok=True)

def get_db():
    conn = sqlite3.connect(DATABASE_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS users (
            uid TEXT PRIMARY KEY,
            displayName TEXT,
            email TEXT UNIQUE,
            photoURL TEXT,
            agencyId TEXT,
            createdAt TEXT,
            firstName TEXT,
            lastName TEXT,
            phone TEXT,
            password TEXT,
            profile_image TEXT,
            role TEXT DEFAULT 'agent'
        )
    ''')
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS agencies (
            id TEXT PRIMARY KEY,
            name TEXT,
            createdById TEXT,
            createdAt TEXT
        )
    ''')
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS portfolios (
            id TEXT PRIMARY KEY,
            agencyId TEXT,
            createdById TEXT,
            createdByName TEXT,
            createdByPhoto TEXT,
            createdAt TEXT,
            title TEXT,
            price REAL,
            type TEXT,
            propertyType TEXT,
            rooms TEXT,
            area REAL,
            city TEXT,
            district TEXT,
            neighborhood TEXT,
            latitude REAL,
            longitude REAL,
            status TEXT,
            heating TEXT,
            age INTEGER,
            floors INTEGER,
            titleStatus TEXT,
            commission REAL,
            notes TEXT,
            imageUrl TEXT,
            sozlesme_tipi TEXT,
            sozlesme_bitis_tarihi TEXT,
            mulk_durumu TEXT,
            tapu_durumu_notlari TEXT
        )
    ''')
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS customers (
            id TEXT PRIMARY KEY,
            agencyId TEXT,
            createdById TEXT,
            createdByName TEXT,
            createdAt TEXT,
            name TEXT,
            phone TEXT,
            email TEXT,
            birthDate TEXT,
            birth_date TEXT,
            type TEXT,
            budget REAL,
            searchLocation TEXT,
            searchRooms TEXT,
            searchPropertyType TEXT,
            notes TEXT,
            finansman_tipi TEXT,
            satin_alma_amaci TEXT,
            yabanci_satis TEXT,
            aciliyet_durumu TEXT,
            sozlesme_tipi TEXT,
            sozlesme_bitis_tarihi TEXT,
            mulk_durumu TEXT,
            tapu_durumu_notlari TEXT
        )
    ''')
    
    # Alter tables to add new columns if they do not exist
    alter_queries = [
        ("users", "firstName", "TEXT"),
        ("users", "lastName", "TEXT"),
        ("users", "phone", "TEXT"),
        ("users", "password", "TEXT"),
        ("users", "profile_image", "TEXT"),
        ("users", "role", "TEXT DEFAULT 'agent'"),
        ("customers", "birth_date", "TEXT")
    ]
    for table, col, col_type in alter_queries:
        try:
            cursor.execute(f"ALTER TABLE {table} ADD COLUMN {col} {col_type}")
        except sqlite3.OperationalError:
            pass # column already exists
            
    # Initialize firstName and lastName for existing users if empty
    try:
        cursor.execute("SELECT uid, displayName, firstName, lastName FROM users")
        users_rows = cursor.fetchall()
        for u in users_rows:
            uid = u["uid"]
            disp = u["displayName"] or ""
            fname = u["firstName"]
            lname = u["lastName"]
            if (fname is None or fname == "") and disp:
                parts = disp.split(" ", 1)
                f = parts[0]
                l = parts[1] if len(parts) > 1 else ""
                cursor.execute("UPDATE users SET firstName = ?, lastName = ? WHERE uid = ?", (f, l, uid))
    except Exception as ex:
        print("Error migrating display names:", ex)
        
    conn.commit()
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS meetings (
            id TEXT PRIMARY KEY,
            agencyId TEXT,
            createdById TEXT,
            createdByName TEXT,
            createdAt TEXT,
            customerId TEXT,
            customerName TEXT,
            title TEXT,
            type TEXT,
            date TEXT,
            time TEXT,
            notes TEXT,
            kanbanStage TEXT
        )
    ''')
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS locations (
            id TEXT PRIMARY KEY,
            name TEXT,
            notes TEXT,
            sqmPriceSale REAL,
            sqmPriceRent REAL,
            competitorNotes TEXT,
            demographics TEXT,
            trends TEXT,
            subNeighborhoods TEXT,
            createdAt TEXT
        )
    ''')
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS todos (
            id TEXT PRIMARY KEY,
            agencyId TEXT,
            task TEXT,
            completed INTEGER DEFAULT 0,
            dueDate TEXT,
            assignedToId TEXT,
            assignedToName TEXT
        )
    ''')
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS activities (
            id TEXT PRIMARY KEY,
            agencyId TEXT,
            userName TEXT,
            userPhoto TEXT,
            action TEXT,
            time TEXT
        )
    ''')
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS deals (
            id TEXT PRIMARY KEY,
            agencyId TEXT,
            createdById TEXT,
            createdByName TEXT,
            createdAt TEXT,
            portfolioId TEXT,
            portfolioTitle TEXT,
            buyerId TEXT,
            buyerName TEXT,
            sellerId TEXT,
            sellerName TEXT,
            agreedPrice REAL,
            buyerInvoiceStatus TEXT,
            sellerInvoiceStatus TEXT,
            checklist TEXT
        )
    ''')
    
    cursor.execute('SELECT COUNT(*) FROM locations')
    if cursor.fetchone()[0] == 0:
        default_locations = [
            {
                "id": "l1",
                "name": "Kadıköy Göztepe",
                "notes": "Ulaşım kolay (Marmaray ve metroya yakın), kentsel dönüşüm sebebiyle yeni binalar yoğunlukta. Prestijli bir bölge.",
                "sqmPriceSale": 95000,
                "sqmPriceRent": 350,
                "competitorNotes": "Bölgede Remax ve Coldwell Banker aktif. Fiyatlar son 6 ayda %15 arttı.",
                "demographics": "A+ Sosyo-ekonomik düzey, emekli ve beyaz yakalı aileler yoğunlukta.",
                "trends": json.dumps([
                    { "year": "2024", "salePrice": 70000, "rentPrice": 240 },
                    { "year": "2025", "salePrice": 85000, "rentPrice": 300 },
                    { "year": "2026", "salePrice": 95000, "rentPrice": 350 }
                ]),
                "subNeighborhoods": json.dumps([]),
                "createdAt": "2026-05-22T00:00:00Z"
            },
            {
                "id": "l2",
                "name": "Suadiye Sahil",
                "notes": "Sahil şeridi çok talep görüyor. Fiyat marjları oldukça yüksek. Yatırım ve oturum için ideal.",
                "sqmPriceSale": 150000,
                "sqmPriceRent": 500,
                "competitorNotes": "Yerel lüks butik emlakçılar çok güçlü. İlan paylaşımları (portföy havuzu) yaygın.",
                "demographics": "Yüksek gelir grubu, nezih profil.",
                "trends": json.dumps([
                    { "year": "2024", "salePrice": 110000, "rentPrice": 380 },
                    { "year": "2025", "salePrice": 135000, "rentPrice": 450 },
                    { "year": "2026", "salePrice": 150000, "rentPrice": 500 }
                ]),
                "subNeighborhoods": json.dumps([]),
                "createdAt": "2026-05-22T00:00:00Z"
            },
            {
                "id": "l3",
                "name": "Kartal",
                "notes": "Kentsel dönüşümün en yoğun olduğu bölgelerden biri. Sahil kesiminde lüks rezidans projeleri, iç kesimlerde ise daha bütçe dostu konutlar yer alıyor. Metro ve Marmaray bağlantısı güçlü.",
                "sqmPriceSale": 42000,
                "sqmPriceRent": 190,
                "competitorNotes": "Bölgede yerel emlak ofisleri ve büyük franchise markalar oldukça agresif pazarlama yapıyor.",
                "demographics": "Orta-üst gelir grubu, yeni evli çiftler ve beyaz yakalı çalışanlar.",
                "trends": json.dumps([
                    { "year": "2024", "salePrice": 32000, "rentPrice": 140 },
                    { "year": "2025", "salePrice": 38000, "rentPrice": 170 },
                    { "year": "2026", "salePrice": 42000, "rentPrice": 190 }
                ]),
                "subNeighborhoods": json.dumps([
                    { "name": "Kordonboyu", "sale": 55000, "rent": 250 },
                    { "name": "Petrol İş", "sale": 40000, "rent": 180 },
                    { "name": "Karlıktepe", "sale": 38000, "rent": 170 },
                    { "name": "Orhantepe", "sale": 44000, "rent": 200 }
                ]),
                "createdAt": "2026-05-22T00:00:00Z"
            },
            {
                "id": "l4",
                "name": "Maltepe",
                "notes": "Sahil yolu, minibüs caddesi ve metro hattı olmak üzere üç ana ulaşım aksına sahip. Kentsel dönüşümle çehresi hızla yenileniyor. Geniş sosyal alanlara ev sahipliği yapıyor.",
                "sqmPriceSale": 48000,
                "sqmPriceRent": 220,
                "competitorNotes": "Maltepe merkezli butik danışmanlık firmaları lüks segmentte pazar payına sahip.",
                "demographics": "Üst-orta gelir grubu, üniversite öğrencileri ve köklü Maltepe sakinleri.",
                "trends": json.dumps([
                    { "year": "2024", "salePrice": 36000, "rentPrice": 160 },
                    { "year": "2025", "salePrice": 43000, "rentPrice": 195 },
                    { "year": "2026", "salePrice": 48000, "rentPrice": 220 }
                ]),
                "subNeighborhoods": json.dumps([
                    { "name": "Küçükyalı", "sale": 62000, "rent": 280 },
                    { "name": "İdealtepe", "sale": 58000, "rent": 260 },
                    { "name": "Altıntepe", "sale": 46000, "rent": 210 },
                    { "name": "Yalı", "sale": 50000, "rent": 230 }
                ]),
                "createdAt": "2026-05-22T00:00:00Z"
            },
            {
                "id": "l5",
                "name": "Pendik",
                "notes": "Sabiha Gökçen Havalimanı ve teknopark yatırımları ile sanayi/teknoloji çalışanlarının uğrak noktası. Uygun fiyatlı konut stoğu barındırırken sahil kesiminde villa ve lüks daire seçenekleri de sunuyor.",
                "sqmPriceSale": 35000,
                "sqmPriceRent": 155,
                "competitorNotes": "Pendik ve Kurtköy bölgelerinde yeni projelerin satış ofisleri doğrudan satış yapıyor.",
                "demographics": "Orta gelir grubu, sanayi ve havacılık sektörü çalışanları, kalabalık aileler.",
                "trends": json.dumps([
                    { "year": "2024", "salePrice": 27000, "rentPrice": 110 },
                    { "year": "2025", "salePrice": 31000, "rentPrice": 135 },
                    { "year": "2026", "salePrice": 35000, "rentPrice": 155 }
                ]),
                "subNeighborhoods": json.dumps([
                    { "name": "Batı", "sale": 45000, "rent": 200 },
                    { "name": "Bahçelievler", "sale": 37000, "rent": 165 },
                    { "name": "Yeni Mahalle", "sale": 34000, "rent": 150 }
                ]),
                "createdAt": "2026-05-22T00:00:00Z"
            }
        ]
        for loc in default_locations:
            cursor.execute('''
                INSERT INTO locations (id, name, notes, sqmPriceSale, sqmPriceRent, competitorNotes, demographics, trends, subNeighborhoods, createdAt)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (loc["id"], loc["name"], loc["notes"], loc["sqmPriceSale"], loc["sqmPriceRent"], loc["competitorNotes"], loc["demographics"], loc["trends"], loc["subNeighborhoods"], loc["createdAt"]))
            
    # Promote sariermusa4137@gmail.com to admin role
    try:
        cursor.execute("UPDATE users SET role = 'admin' WHERE email = 'sariermusa4137@gmail.com'")
        conn.commit()
    except Exception as e:
        print(f"Error promoting admin: {e}", flush=True)
        
    # Create and populate rol_yetkileri table
    try:
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS rol_yetkileri (
                role TEXT PRIMARY KEY,
                can_delete_portfolio INTEGER DEFAULT 1,
                can_edit_customer INTEGER DEFAULT 1,
                can_view_all_agency INTEGER DEFAULT 1,
                can_view_reports INTEGER DEFAULT 1
            )
        ''')
        
        # Add new columns if table existed
        alter_queries_yetki = [
            ("rol_yetkileri", "can_view_reports", "INTEGER DEFAULT 1")
        ]
        for t, col, c_type in alter_queries_yetki:
            try:
                cursor.execute(f"ALTER TABLE {t} ADD COLUMN {col} {c_type}")
            except sqlite3.OperationalError:
                pass
                
        cursor.execute("SELECT COUNT(*) FROM rol_yetkileri WHERE role = 'admin'")
        if cursor.fetchone()[0] == 0:
            cursor.execute("INSERT INTO rol_yetkileri (role, can_delete_portfolio, can_edit_customer, can_view_all_agency, can_view_reports) VALUES ('admin', 1, 1, 1, 1)")
        cursor.execute("SELECT COUNT(*) FROM rol_yetkileri WHERE role = 'agent'")
        if cursor.fetchone()[0] == 0:
            cursor.execute("INSERT INTO rol_yetkileri (role, can_delete_portfolio, can_edit_customer, can_view_all_agency, can_view_reports) VALUES ('agent', 1, 1, 1, 0)")
        cursor.execute("SELECT COUNT(*) FROM rol_yetkileri WHERE role = 'assistant'")
        if cursor.fetchone()[0] == 0:
            cursor.execute("INSERT INTO rol_yetkileri (role, can_delete_portfolio, can_edit_customer, can_view_all_agency, can_view_reports) VALUES ('assistant', 0, 1, 1, 0)")
        conn.commit()
    except Exception as e:
        print(f"Error initializing rol_yetkileri table: {e}", flush=True)
        
    conn.close()

# Initialize database on app startup (runs under Gunicorn and development server)
with app.app_context():
    init_db()
    # Temporary password reset migration hook
    try:
        conn = get_db()
        cursor = conn.cursor()
        email_to_reset = 'sariermusa4137@gmail.com'
        cursor.execute('SELECT * FROM users WHERE email = ?', (email_to_reset,))
        user = cursor.fetchone()
        if user:
            new_hash = generate_password_hash('41374137', method='pbkdf2:sha256')
            cursor.execute('UPDATE users SET password = ? WHERE email = ?', (new_hash, email_to_reset))
            conn.commit()
            print(f"Password reset migration executed for {email_to_reset}", flush=True)
        conn.close()
    except Exception as e:
        print(f"Error in password reset migration: {e}", flush=True)

@app.route('/api/auth/register', methods=['POST'])
def auth_register():
    try:
        data = request.get_json() or {}
        first_name = data.get('firstName', '').strip()
        last_name = data.get('lastName', '').strip()
        email = data.get('email', '').strip().lower()
        password = data.get('password', '')
        
        if not first_name or not email or not password:
            return {"error": "Ad, E-posta ve Şifre alanları zorunludur."}, 400
            
        conn = get_db()
        cursor = conn.cursor()
        
        # Check if email is already registered
        cursor.execute('SELECT * FROM users WHERE email = ?', (email,))
        if cursor.fetchone():
            conn.close()
            return {"error": "Bu e-posta adresi zaten kayıtlıdır."}, 400
            
        # Hash password and compute deterministic user properties
        hashed_password = generate_password_hash(password, method='pbkdf2:sha256')
        uid = str(uuid.uuid5(uuid.NAMESPACE_DNS, email))
        created_at = datetime.datetime.now().isoformat()
        display_name = f"{first_name} {last_name}".strip()
        
        # Automatically assign their own 'uid' as personal agencyId and create agency
        agency_id = uid
        agency_name = f"{display_name} (Bireysel)"
        
        cursor.execute('''
            INSERT OR IGNORE INTO agencies (id, name, createdById, createdAt)
            VALUES (?, ?, ?, ?)
        ''', (agency_id, agency_name, uid, created_at))
        
        cursor.execute('''
            INSERT INTO users (uid, displayName, email, photoURL, agencyId, createdAt, firstName, lastName, phone, password, profile_image)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (uid, display_name, email, "", agency_id, created_at, first_name, last_name, "", hashed_password, ""))
        
        conn.commit()
        
        cursor.execute('SELECT * FROM users WHERE uid = ?', (uid,))
        user_data = dict(cursor.fetchone())
        
        cursor.execute('SELECT * FROM agencies WHERE id = ?', (agency_id,))
        agency_data = dict(cursor.fetchone())
        
        conn.close()
        
        # Start persistent session
        session['user_id'] = uid
        session.permanent = True
        
        return {"success": True, "user": user_data, "agency": agency_data}
    except Exception as e:
        return {"error": str(e)}, 500

@app.route('/api/auth/login', methods=['POST'])
def auth_login():
    try:
        data = request.get_json() or {}
        email = data.get('email', '').strip().lower()
        password = data.get('password', '')
        
        if not email or not password:
            return {"error": "E-posta ve Şifre alanları zorunludur."}, 400
            
        conn = get_db()
        cursor = conn.cursor()
        
        cursor.execute('SELECT * FROM users WHERE email = ?', (email,))
        user = cursor.fetchone()
        
        if not user:
            conn.close()
            return {"error": "E-posta adresi veya şifre hatalı."}, 401
            
        user_data = dict(user)
        
        # Check password hash
        hashed_pwd = user_data.get('password')
        if not hashed_pwd or not check_password_hash(hashed_pwd, password):
            conn.close()
            return {"error": "E-posta adresi veya şifre hatalı."}, 401
            
        agency_data = None
        if user_data['agencyId']:
            cursor.execute('SELECT * FROM agencies WHERE id = ?', (user_data['agencyId'],))
            agency = cursor.fetchone()
            if agency:
                agency_data = dict(agency)
                
        conn.close()
        
        # Set session cookie
        session['user_id'] = user_data['uid']
        session.permanent = True
        
        return {"user": user_data, "agency": agency_data}
    except Exception as e:
        return {"error": str(e)}, 500

@app.route('/api/auth/status', methods=['GET'])
def auth_status():
    try:
        user_id = session.get('user_id')
        if not user_id:
            user_id = request.args.get('userId')
            if not user_id:
                return {"error": "Oturum bulunamadı. Lütfen giriş yapın."}, 401
            # Auto-restore session from valid localStorage identifier
            session['user_id'] = user_id
            session.permanent = True
            
        conn = get_db()
        cursor = conn.cursor()
        
        cursor.execute('SELECT * FROM users WHERE uid = ?', (user_id,))
        user = cursor.fetchone()
        
        if not user:
            conn.close()
            # Clear invalid session
            session.clear()
            return {"error": "Kullanıcı bulunamadı."}, 404
            
        user_data = dict(user)
        
        # Ensure user has a personal workspace if they don't have an agencyId
        if not user_data.get('agencyId'):
            user_data['agencyId'] = user_id
            cursor.execute('UPDATE users SET agencyId = ? WHERE uid = ?', (user_id, user_id))
            conn.commit()
            
        # Ensure agency exists
        cursor.execute('SELECT * FROM agencies WHERE id = ?', (user_data['agencyId'],))
        agency = cursor.fetchone()
        if not agency:
            # Create a default personal agency
            agency_name = f"{user_data['displayName']} (Bireysel)"
            cursor.execute('''
                INSERT INTO agencies (id, name, createdById, createdAt)
                VALUES (?, ?, ?, ?)
            ''', (user_data['agencyId'], agency_name, user_id, datetime.datetime.now().isoformat()))
            conn.commit()
            
            cursor.execute('SELECT * FROM agencies WHERE id = ?', (user_data['agencyId'],))
            agency = cursor.fetchone()
            
        agency_data = dict(agency) if agency else None
        conn.close()
        return {"user": user_data, "agency": agency_data}
    except Exception as e:
        return {"error": str(e)}, 500

@app.route('/api/agency/create', methods=['POST'])
@login_required
def agency_create():
    try:
        data = request.get_json() or {}
        user_id = data.get('userId')
        agency_name = data.get('name', '').strip()
        
        if not user_id or not agency_name:
            return {"error": "Kullanıcı ID ve Acente adı gereklidir."}, 400
            
        conn = get_db()
        cursor = conn.cursor()
        
        cursor.execute('SELECT * FROM users WHERE uid = ?', (user_id,))
        user = cursor.fetchone()
        if not user:
            conn.close()
            return {"error": "Kullanıcı bulunamadı."}, 404
            
        agency_code = ''.join(random.choices(string.ascii_uppercase + string.digits, k=6))
        created_at = datetime.datetime.now().isoformat()
        
        cursor.execute('''
            INSERT INTO agencies (id, name, createdById, createdAt)
            VALUES (?, ?, ?, ?)
        ''', (agency_code, agency_name, user_id, created_at))
        
        cursor.execute('UPDATE users SET agencyId = ? WHERE uid = ?', (agency_code, user_id))
        conn.commit()
        
        cursor.execute('SELECT * FROM agencies WHERE id = ?', (agency_code,))
        agency_data = dict(cursor.fetchone())
        
        conn.close()
        return {"agency": agency_data}
    except Exception as e:
        return {"error": str(e)}, 500

@app.route('/api/agency/join', methods=['POST'])
@login_required
def agency_join():
    try:
        data = request.get_json() or {}
        user_id = data.get('userId')
        agency_code = data.get('agencyCode', '').strip().upper()
        
        if not user_id or not agency_code:
            return {"error": "Kullanıcı ID ve Acente kodu gereklidir."}, 400
            
        conn = get_db()
        cursor = conn.cursor()
        
        cursor.execute('SELECT * FROM users WHERE uid = ?', (user_id,))
        user = cursor.fetchone()
        if not user:
            conn.close()
            return {"error": "Kullanıcı bulunamadı."}, 404
            
        cursor.execute('SELECT * FROM agencies WHERE id = ?', (agency_code,))
        agency = cursor.fetchone()
        if not agency:
            conn.close()
            return {"error": "Acente bulunamadı. Lütfen kodu kontrol edin."}, 404
            
        agency_data = dict(agency)
        
        cursor.execute('UPDATE users SET agencyId = ? WHERE uid = ?', (agency_code, user_id))
        conn.commit()
        
        conn.close()
        return {"agency": agency_data}
    except Exception as e:
        return {"error": str(e)}, 500

@app.route('/api/agency/update', methods=['POST'])
@login_required
def agency_update():
    try:
        data = request.get_json() or {}
        agency_id = data.get('agencyId')
        name = data.get('name', '').strip()
        
        if not agency_id or not name:
            return {"error": "Acente ID ve yeni isim gereklidir."}, 400
            
        conn = get_db()
        cursor = conn.cursor()
        
        # Verify user is member of this agency
        cursor.execute('SELECT agencyId FROM users WHERE uid = ?', (session.get('user_id'),))
        user_row = cursor.fetchone()
        if not user_row or user_row['agencyId'] != agency_id:
            conn.close()
            return {"error": "Bu acenteyi güncelleme yetkiniz yok."}, 403
            
        cursor.execute('UPDATE agencies SET name = ? WHERE id = ?', (name, agency_id))
        conn.commit()
        
        cursor.execute('SELECT * FROM agencies WHERE id = ?', (agency_id,))
        agency_data = dict(cursor.fetchone())
        conn.close()
        
        return {"success": True, "agency": agency_data}
    except Exception as e:
        return {"error": str(e)}, 500

@app.route('/api/data', methods=['GET'])
@login_required
def get_all_data():
    try:
        agency_id = request.args.get('agencyId')
        if not agency_id:
            return {"error": "agencyId parametresi gereklidir."}, 400
            
        conn = get_db()
        cursor = conn.cursor()
        
        # Get permissions
        current_user_id = session.get('user_id')
        cursor.execute('SELECT role FROM users WHERE uid = ?', (current_user_id,))
        user_row = cursor.fetchone()
        role = user_row['role'] if user_row else 'agent'
        if not role:
            role = 'agent'
            
        cursor.execute('SELECT can_view_all_agency FROM rol_yetkileri WHERE role = ?', (role,))
        perm_row = cursor.fetchone()
        can_view_all_agency = perm_row['can_view_all_agency'] if perm_row else 1
        
        tables = ['portfolios', 'customers', 'meetings', 'todos', 'activities', 'deals']
        response_data = {}
        
        for table in tables:
            query = f'SELECT * FROM {table} WHERE agencyId = ?'
            params = [agency_id]
            
            if not can_view_all_agency:
                if table in ['portfolios', 'customers', 'meetings', 'deals']:
                    query += ' AND createdById = ?'
                    params.append(current_user_id)
                elif table == 'todos':
                    query += ' AND assignedToId = ?'
                    params.append(current_user_id)
            
            query += ' ORDER BY rowid DESC'
            cursor.execute(query, params)
            rows = cursor.fetchall()
            list_data = []
            for r in rows:
                item = dict(r)
                if table == 'todos':
                    item['completed'] = bool(item['completed'])
                elif table == 'deals':
                    try:
                        item['checklist'] = json.loads(item['checklist']) if item.get('checklist') else []
                    except:
                        item['checklist'] = []
                list_data.append(item)
                
            if table == 'meetings':
                try:
                    cust_query = 'SELECT id, name, birthDate, birth_date FROM customers WHERE agencyId = ?'
                    cust_params = [agency_id]
                    if not can_view_all_agency:
                        cust_query += ' AND createdById = ?'
                        cust_params.append(current_user_id)
                    cursor.execute(cust_query, cust_params)
                    customers_rows = cursor.fetchall()
                    today_year = datetime.date.today().year
                    for c_row in customers_rows:
                        c_id = c_row['id']
                        c_name = c_row['name']
                        b_str = c_row['birth_date'] or c_row['birthDate']
                        if b_str:
                            parts = b_str.split('-')
                            if len(parts) >= 2:
                                try:
                                    if len(parts) == 3:
                                        m_val = int(parts[1])
                                        d_val = int(parts[2])
                                    else:
                                        m_val = int(parts[0])
                                        d_val = int(parts[1])
                                    
                                    for yr in [today_year - 1, today_year, today_year + 1]:
                                        list_data.append({
                                            "id": f"birthday-{c_id}-{yr}",
                                            "agencyId": agency_id,
                                            "createdById": "system",
                                            "createdByName": "Sistem",
                                            "createdAt": "",
                                            "customerId": c_id,
                                            "customerName": c_name,
                                            "title": f"🎂 Doğum Günü: {c_name}",
                                            "type": "Doğum Günü",
                                            "date": f"{yr}-{m_val:02d}-{d_val:02d}",
                                            "time": "09:00",
                                            "notes": f"{c_name} Doğum Günü",
                                            "kanbanStage": ""
                                        })
                                except (ValueError, IndexError):
                                    pass
                except Exception as ex:
                    print("Error injecting birthdays:", ex)
                    
            response_data[table] = list_data
            
        cursor.execute('SELECT * FROM locations')
        rows = cursor.fetchall()
        loc_data = []
        for r in rows:
            item = dict(r)
            try:
                item['trends'] = json.loads(item['trends']) if item.get('trends') else []
            except:
                item['trends'] = []
            try:
                item['subNeighborhoods'] = json.loads(item['subNeighborhoods']) if item.get('subNeighborhoods') else []
            except:
                item['subNeighborhoods'] = []
            loc_data.append(item)
        response_data['locations'] = loc_data
        
        conn.close()
        return response_data
    except Exception as e:
        return {"error": str(e)}, 500

@app.route('/api/data/<collection>', methods=['POST'])
@login_required
def create_data_record(collection):
    try:
        data = request.get_json() or {}
        
        if collection == 'customers':
            if 'birthDate' in data and 'birth_date' not in data:
                data['birth_date'] = data['birthDate']
            elif 'birth_date' in data and 'birthDate' not in data:
                data['birthDate'] = data['birth_date']
        
        allowed_collections = ['portfolios', 'customers', 'meetings', 'todos', 'activities', 'deals']
        if collection not in allowed_collections:
            return {"error": "Geçersiz koleksiyon adı."}, 400
            
        conn = get_db()
        cursor = conn.cursor()
        
        record_id = data.get('id') or str(uuid.uuid4())
        data['id'] = record_id
        
        cursor.execute(f'PRAGMA table_info({collection})')
        columns = [col[1] for col in cursor.fetchall()]
        
        cols_to_insert = []
        vals_to_insert = []
        
        for col in columns:
            if col in data:
                val = data[col]
                if col in ['trends', 'subNeighborhoods', 'checklist'] and not isinstance(val, str):
                    val = json.dumps(val)
                elif col == 'completed':
                    val = 1 if val else 0
                cols_to_insert.append(col)
                vals_to_insert.append(val)
                
        if 'id' not in cols_to_insert:
            cols_to_insert.append('id')
            vals_to_insert.append(record_id)
            
        placeholders = ', '.join(['?' for _ in vals_to_insert])
        query = f"INSERT INTO {collection} ({', '.join(cols_to_insert)}) VALUES ({placeholders})"
        cursor.execute(query, vals_to_insert)
        conn.commit()
        
        cursor.execute(f"SELECT * FROM {collection} WHERE id = ?", (record_id,))
        row = cursor.fetchone()
        inserted_data = dict(row)
        if collection == 'todos':
            inserted_data['completed'] = bool(inserted_data['completed'])
        elif collection == 'deals':
            try:
                inserted_data['checklist'] = json.loads(inserted_data['checklist']) if inserted_data.get('checklist') else []
            except:
                inserted_data['checklist'] = []
                
        conn.close()
        return inserted_data
    except Exception as e:
        import traceback
        traceback.print_exc()
        return {"error": str(e)}, 500

@app.route('/api/data/<collection>/<id>', methods=['PUT'])
@login_required
def update_data_record(collection, id):
    try:
        data = request.get_json() or {}
        
        if collection == 'customers':
            if 'birthDate' in data and 'birth_date' not in data:
                data['birth_date'] = data['birthDate']
            elif 'birth_date' in data and 'birthDate' not in data:
                data['birthDate'] = data['birth_date']
        
        allowed_collections = ['portfolios', 'customers', 'meetings', 'todos', 'activities', 'deals']
        if collection not in allowed_collections:
            return {"error": "Geçersiz koleksiyon adı."}, 400
            
        conn = get_db()
        cursor = conn.cursor()
        
        # Verify permissions
        current_user_id = session.get('user_id')
        cursor.execute('SELECT role FROM users WHERE uid = ?', (current_user_id,))
        user_row = cursor.fetchone()
        role = user_row['role'] if user_row else 'agent'
        if not role:
            role = 'agent'
            
        if collection == 'customers':
            cursor.execute('SELECT can_edit_customer FROM rol_yetkileri WHERE role = ?', (role,))
            perm_row = cursor.fetchone()
            can_edit_customer = perm_row['can_edit_customer'] if perm_row else 1
            if not can_edit_customer:
                conn.close()
                return {"error": "Müşteri düzenleme yetkiniz bulunmamaktadır."}, 403
        
        cursor.execute(f'PRAGMA table_info({collection})')
        columns = [col[1] for col in cursor.fetchall()]
        
        updates = []
        vals = []
        for col in columns:
            if col in data and col != 'id':
                val = data[col]
                if col in ['trends', 'subNeighborhoods', 'checklist'] and not isinstance(val, str):
                    val = json.dumps(val)
                elif col == 'completed':
                    val = 1 if val else 0
                updates.append(f"{col} = ?")
                vals.append(val)
                
        if not updates:
            conn.close()
            return {"error": "Güncellenecek alan gönderilmedi."}, 400
            
        vals.append(id)
        query = f"UPDATE {collection} SET {', '.join(updates)} WHERE id = ?"
        cursor.execute(query, vals)
        conn.commit()
        
        cursor.execute(f"SELECT * FROM {collection} WHERE id = ?", (id,))
        row = cursor.fetchone()
        updated_data = dict(row)
        if collection == 'todos':
            updated_data['completed'] = bool(updated_data['completed'])
        elif collection == 'deals':
            try:
                updated_data['checklist'] = json.loads(updated_data['checklist']) if updated_data.get('checklist') else []
            except:
                updated_data['checklist'] = []
                
        conn.close()
        return updated_data
    except Exception as e:
        import traceback
        traceback.print_exc()
        return {"error": str(e)}, 500

@app.route('/api/data/<collection>/<id>', methods=['DELETE'])
@login_required
def delete_data_record(collection, id):
    try:
        allowed_collections = ['portfolios', 'customers', 'meetings', 'todos', 'activities', 'deals']
        if collection not in allowed_collections:
            return {"error": "Geçersiz koleksiyon adı."}, 400
            
        conn = get_db()
        cursor = conn.cursor()
        
        # Verify permissions
        current_user_id = session.get('user_id')
        cursor.execute('SELECT role FROM users WHERE uid = ?', (current_user_id,))
        user_row = cursor.fetchone()
        role = user_row['role'] if user_row else 'agent'
        if not role:
            role = 'agent'
            
        if collection == 'portfolios':
            cursor.execute('SELECT can_delete_portfolio FROM rol_yetkileri WHERE role = ?', (role,))
            perm_row = cursor.fetchone()
            can_delete_portfolio = perm_row['can_delete_portfolio'] if perm_row else 1
            if not can_delete_portfolio:
                conn.close()
                return {"error": "Portföy silme yetkiniz bulunmamaktadır."}, 403
        
        cursor.execute(f"DELETE FROM {collection} WHERE id = ?", (id,))
        conn.commit()
        
        conn.close()
        return {"success": True}
    except Exception as e:
        return {"error": str(e)}, 500

@app.route('/api/profile/update', methods=['POST'])
@login_required
def profile_update():
    try:
        data = request.get_json() or {}
        uid = data.get('uid')
        first_name = data.get('firstName', '').strip()
        last_name = data.get('lastName', '').strip()
        phone = data.get('phone', '').strip()
        email = data.get('email', '').strip().lower()
        password = data.get('password', '')
        
        if not uid or not first_name or not email:
            return {"error": "Kullanıcı ID, Ad ve E-posta alanları zorunludur."}, 400
            
        conn = get_db()
        cursor = conn.cursor()
        
        # Check email uniqueness
        cursor.execute('SELECT uid FROM users WHERE email = ? AND uid != ?', (email, uid))
        if cursor.fetchone():
            conn.close()
            return {"error": "Bu e-posta adresi başka bir kullanıcı tarafından kullanılmaktadır."}, 400
            
        display_name = f"{first_name} {last_name}".strip()
        
        query = '''
            UPDATE users 
            SET firstName = ?, lastName = ?, displayName = ?, phone = ?, email = ?
        '''
        params = [first_name, last_name, display_name, phone, email]
        
        if password:
            query += ", password = ?"
            params.append(password)
            
        query += " WHERE uid = ?"
        params.append(uid)
        
        cursor.execute(query, params)
        conn.commit()
        
        # Get updated user data
        cursor.execute('SELECT * FROM users WHERE uid = ?', (uid,))
        user_row = cursor.fetchone()
        user_data = dict(user_row)
        conn.close()
        
        return {"success": True, "user": user_data}
    except Exception as e:
        return {"error": str(e)}, 500

@app.route('/api/profile/upload', methods=['POST'])
@login_required
def profile_upload():
    try:
        if 'file' not in request.files:
            return {"error": "Yüklenecek dosya bulunamadı."}, 400
            
        file = request.files['file']
        uid = request.form.get('uid')
        
        if not uid:
            return {"error": "Kullanıcı ID gereklidir."}, 400
            
        if file.filename == '':
            return {"error": "Dosya seçilmedi."}, 400
            
        if file:
            allowed_extensions = {'png', 'jpg', 'jpeg', 'gif', 'webp'}
            filename = file.filename
            ext = filename.rsplit('.', 1)[1].lower() if '.' in filename else ''
            if ext not in allowed_extensions:
                return {"error": "Sadece resim dosyaları (.png, .jpg, .jpeg, .gif, .webp) yüklenebilir."}, 400
                
            upload_folder = os.path.join(app.config['UPLOAD_FOLDER'], 'profiles')
            os.makedirs(upload_folder, exist_ok=True)
            
            unique_filename = f"{uid}_{uuid.uuid4().hex[:8]}.{ext}"
            file_path = os.path.join(upload_folder, unique_filename)
            file.save(file_path)
            
            relative_url = f"/uploads/profiles/{unique_filename}"
            
            conn = get_db()
            cursor = conn.cursor()
            cursor.execute('UPDATE users SET profile_image = ?, photoURL = ? WHERE uid = ?', (relative_url, relative_url, uid))
            conn.commit()
            conn.close()
            return {"success": True, "profile_image": relative_url}
    except Exception as e:
        return {"error": str(e)}, 500

@app.route('/api/users', methods=['GET'])
@login_required
def get_users():
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT u.uid, u.displayName, u.email, u.photoURL, u.agencyId, u.createdAt, 
                   u.firstName, u.lastName, u.phone, u.profile_image, u.role,
                   a.createdById as agencyOwnerId, a.name as agencyName
            FROM users u
            LEFT JOIN agencies a ON u.agencyId = a.id
            ORDER BY u.createdAt DESC
        ''')
        rows = cursor.fetchall()
        users_list = []
        for r in rows:
            item = dict(r)
            # Default empty role to 'agent'
            if not item.get('role'):
                item['role'] = 'agent'
            
            # Keep agencyOwnerId to calculate if owner, but keep it clean
            is_owner = item.get('uid') == item.get('agencyOwnerId')
            item['isAgencyOwner'] = is_owner
            if 'agencyOwnerId' in item:
                del item['agencyOwnerId']
                
            users_list.append(item)
            
        conn.close()
        return jsonify(users_list)
    except Exception as e:
        return {"error": str(e)}, 500

@app.route('/api/users/update-role', methods=['POST'])
@login_required
def update_user_role():
    try:
        current_user_id = session.get('user_id')
        data = request.get_json() or {}
        target_user_id = data.get('userId')
        new_role = data.get('newRole')
        
        if not target_user_id or not new_role:
            return {"error": "userId ve newRole alanları zorunludur."}, 400
            
        if new_role not in ['admin', 'agent', 'assistant']:
            return {"error": "Geçersiz rol belirtildi."}, 400
            
        conn = get_db()
        cursor = conn.cursor()
        
        # Verify if current user is admin
        cursor.execute('SELECT role FROM users WHERE uid = ?', (current_user_id,))
        current_user = cursor.fetchone()
        if not current_user or current_user['role'] != 'admin':
            conn.close()
            return {"error": "Yetkisiz işlem. Yalnızca yöneticiler rol değiştirebilir."}, 403
            
        # Update user role
        cursor.execute('UPDATE users SET role = ? WHERE uid = ?', (new_role, target_user_id))
        conn.commit()
        conn.close()
        
        return {"success": True}
    except Exception as e:
        return {"error": str(e)}, 500

@app.route('/api/users/delete', methods=['POST'])
@login_required
def delete_user():
    try:
        current_user_id = session.get('user_id')
        data = request.get_json() or {}
        target_user_id = data.get('userId')
        
        if not target_user_id:
            return {"error": "userId alanı zorunludur."}, 400
            
        if current_user_id == target_user_id:
            return {"error": "Yöneticinin kendi kendini silmesi engellenmiştir."}, 400
            
        conn = get_db()
        cursor = conn.cursor()
        
        # Verify if current user is admin
        cursor.execute('SELECT role FROM users WHERE uid = ?', (current_user_id,))
        current_user = cursor.fetchone()
        if not current_user or current_user['role'] != 'admin':
            conn.close()
            return {"error": "Yetkisiz işlem. Yalnızca yöneticiler kullanıcı silebilir."}, 403
            
        # Delete user
        cursor.execute('DELETE FROM users WHERE uid = ?', (target_user_id,))
        conn.commit()
        conn.close()
        
        return {"success": True}
    except Exception as e:
        return {"error": str(e)}, 500

@app.route('/api/users/permissions', methods=['GET'])
@login_required
def get_role_permissions():
    try:
        current_user_id = session.get('user_id')
        conn = get_db()
        cursor = conn.cursor()
        
        # Verify if current user is admin
        cursor.execute('SELECT role FROM users WHERE uid = ?', (current_user_id,))
        current_user = cursor.fetchone()
        if not current_user or current_user['role'] != 'admin':
            conn.close()
            return {"error": "Yetkisiz işlem. Yalnızca yöneticiler yetki ayarlarına erişebilir."}, 403
            
        cursor.execute('SELECT * FROM rol_yetkileri')
        rows = cursor.fetchall()
        conn.close()
        
        permissions_dict = {}
        for row in rows:
            permissions_dict[row['role']] = {
                "can_delete_portfolio": bool(row['can_delete_portfolio']),
                "can_edit_customer": bool(row['can_edit_customer']),
                "can_view_all_agency": bool(row['can_view_all_agency']),
                "can_view_reports": bool(row['can_view_reports'])
            }
        return jsonify(permissions_dict)
    except Exception as e:
        return {"error": str(e)}, 500

@app.route('/api/users/update-permissions', methods=['POST'])
@login_required
def update_role_permissions():
    try:
        current_user_id = session.get('user_id')
        data = request.get_json() or {}
        
        conn = get_db()
        cursor = conn.cursor()
        
        # Verify if current user is admin
        cursor.execute('SELECT role FROM users WHERE uid = ?', (current_user_id,))
        current_user = cursor.fetchone()
        if not current_user or current_user['role'] != 'admin':
            conn.close()
            return {"error": "Yetkisiz işlem. Yalnızca yöneticiler yetki değiştirebilir."}, 403
            
        for role in ['admin', 'agent', 'assistant']:
            role_data = data.get(role, {})
            can_delete_portfolio = 1 if role_data.get('can_delete_portfolio') else 0
            can_edit_customer = 1 if role_data.get('can_edit_customer') else 0
            can_view_all_agency = 1 if role_data.get('can_view_all_agency') else 0
            can_view_reports = 1 if role_data.get('can_view_reports') else 0
            
            cursor.execute('''
                INSERT OR REPLACE INTO rol_yetkileri (role, can_delete_portfolio, can_edit_customer, can_view_all_agency, can_view_reports)
                VALUES (?, ?, ?, ?, ?)
            ''', (role, can_delete_portfolio, can_edit_customer, can_view_all_agency, can_view_reports))
            
        conn.commit()
        conn.close()
        return {"success": True}
    except Exception as e:
        return {"error": str(e)}, 500

@app.route('/api/reports/ciro', methods=['GET'])
@login_required
@roles_accepted('can_view_reports')
def get_ciro_report():
    try:
        agency_id = request.args.get('agencyId')
        if not agency_id:
            return {"error": "agencyId parametresi gereklidir."}, 400
            
        conn = get_db()
        cursor = conn.cursor()
        
        # Calculate from database
        cursor.execute("SELECT SUM(agreedPrice) FROM deals WHERE agencyId = ?", (agency_id,))
        total_deals_price = cursor.fetchone()[0] or 0
        
        # Calculate active listing values
        cursor.execute("SELECT SUM(price) FROM portfolios WHERE agencyId = ?", (agency_id,))
        active_listings_val = cursor.fetchone()[0] or 0
        
        conn.close()
        
        # Format or add some historical data
        return jsonify({
            "totalRevenue": total_deals_price,
            "commissionEarned": total_deals_price * 0.04, # 4% total commission (2% buyer + 2% seller)
            "activeListingsValue": active_listings_val,
            "monthlyRevenue": [180000, 240000, 310000, 290000, 420000, total_deals_price * 0.04 if total_deals_price > 0 else 380000],
            "commissionGoal": 1000000,
            "topPerformers": [
                {"name": "Musa Sarıer", "dealsCount": 4, "comm": 120000},
                {"name": "Canan Demir", "dealsCount": 2, "comm": 60000}
            ]
        })
    except Exception as e:
        return {"error": str(e)}, 500

if __name__ == '__main__':
    init_db()
    port = int(os.environ.get('PORT', 8000))
    print(f"Starting server on port {port}...")
    app.run(host='0.0.0.0', port=port, debug=True)
