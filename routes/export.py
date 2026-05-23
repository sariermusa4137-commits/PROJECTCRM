"""
PROJECTCRM - routes/export.py
Excel dışa aktarma rotası (openpyxl ile profesyonel formatlama).
"""

import io
from flask import Blueprint, request, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from auth_middleware import login_required

export_bp = Blueprint('export', __name__)

_FONT_FAMILY = "Segoe UI"


def _build_styles():
    """Excel stilleri sözlüğü oluşturur."""
    thin_side = Side(style='thin', color='CBD5E1')
    thin_border = Border(
        left=thin_side, right=thin_side, top=thin_side, bottom=thin_side
    )
    return {
        "title_font": Font(name=_FONT_FAMILY, size=14, bold=True, color="1E293B"),
        "section_font": Font(name=_FONT_FAMILY, size=12, bold=True, color="0F172A"),
        "header_font": Font(name=_FONT_FAMILY, size=10, bold=True, color="FFFFFF"),
        "data_font": Font(name=_FONT_FAMILY, size=10, color="334155"),
        "header_fill": PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid"),
        "header_buyer_fill": PatternFill(start_color="047857", end_color="047857", fill_type="solid"),
        "header_seller_fill": PatternFill(start_color="475569", end_color="475569", fill_type="solid"),
        "border": thin_border,
        "align_center": Alignment(horizontal='center', vertical='center', wrap_text=True),
        "align_left": Alignment(horizontal='left', vertical='center', wrap_text=True),
        "align_right": Alignment(horizontal='right', vertical='center', wrap_text=True),
    }


def _autofit_columns(ws, skip_rows=None):
    """Sütun genişliklerini içeriğe göre otomatik ayarlar."""
    skip_rows = skip_rows or set()
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row in skip_rows or cell.value is None:
                continue
            val_str = (f"₺{int(cell.value):,}" if cell.number_format == '₺#,##0'
                       else str(cell.value))
            max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)


@export_bp.route('/api/export', methods=['POST'])
@login_required
def export_excel():
    try:
        data = request.get_json() or {}
        customers = data.get('customers', [])
        portfolios = data.get('portfolios', [])

        buyers = [c for c in customers if c.get('type') == 'Alıcı']
        sellers = [c for c in customers if c.get('type') == 'Satıcı']

        wb = Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)
        s = _build_styles()

        # ------------------------------------------------------------------ #
        # Sheet 1: Alıcı Takip Listesi                                        #
        # ------------------------------------------------------------------ #
        ws1 = wb.create_sheet(title="Alıcı Takip Listesi")
        ws1.views.sheetView[0].showGridLines = True
        ws1.cell(row=1, column=1, value="ALICI MÜŞTERİ TAKİP LİSTESİ").font = s["title_font"]
        ws1.row_dimensions[1].height = 30

        headers_buyer = [
            "Müşteri Adı", "Telefon", "E-posta", "Doğum Tarihi",
            "Bütçe", "Hedef Bölge", "Aranan Oda", "Konut Tipi",
            "Finansman Tipi", "Satın Alma Amacı", "Yabancı Satış", "Aciliyet Durumu",
            "Danışman", "Notlar"
        ]
        for col_num, header in enumerate(headers_buyer, 1):
            cell = ws1.cell(row=3, column=col_num, value=header)
            cell.font = s["header_font"]
            cell.fill = s["header_buyer_fill"]
            cell.alignment = s["align_center"]
            cell.border = s["border"]
        ws1.row_dimensions[3].height = 26

        for row_idx, b in enumerate(buyers, 4):
            budget_val = 0
            try:
                budget_val = float(b.get('budget', 0))
            except (ValueError, TypeError):
                pass
            row_data = [
                (b.get('name', ''), s["align_left"]),
                (b.get('phone', ''), s["align_center"]),
                (b.get('email', ''), s["align_left"]),
                (b.get('birthDate', ''), s["align_center"]),
                (budget_val, s["align_right"]),
                (b.get('searchLocation', ''), s["align_left"]),
                (b.get('searchRooms', ''), s["align_center"]),
                (b.get('searchPropertyType', ''), s["align_left"]),
                (b.get('finansman_tipi', ''), s["align_center"]),
                (b.get('satin_alma_amaci', ''), s["align_center"]),
                (b.get('yabanci_satis', ''), s["align_center"]),
                (b.get('aciliyet_durumu', ''), s["align_center"]),
                (b.get('createdByName', ''), s["align_left"]),
                (b.get('notes', ''), s["align_left"]),
            ]
            for col_num, (val, align) in enumerate(row_data, 1):
                cell = ws1.cell(row=row_idx, column=col_num, value=val)
                cell.font = s["data_font"]
                cell.alignment = align
                cell.border = s["border"]
                if col_num == 5:
                    cell.number_format = '₺#,##0'
            ws1.row_dimensions[row_idx].height = 20

        _autofit_columns(ws1, skip_rows={1, 2})

        # ------------------------------------------------------------------ #
        # Sheet 2: Satıcı & Portföy Listesi                                   #
        # ------------------------------------------------------------------ #
        ws2 = wb.create_sheet(title="Satıcı & Portföy Listesi")
        ws2.views.sheetView[0].showGridLines = True
        ws2.cell(row=1, column=1, value="PORTFÖY LİSTESİ").font = s["section_font"]
        ws2.row_dimensions[1].height = 24

        headers_portfolio = [
            "Portföy Başlığı", "Fiyat", "İşlem Tipi", "Konut Tipi", "Oda Sayısı",
            "Alan (m²)", "İlçe", "Mahalle", "Isınma", "Bina Yaşı", "Durum",
            "Sözleşme Tipi", "Sözleşme Bitiş Tarihi", "Mülk Durumu", "Tapu Notları", "Danışman"
        ]
        for col_num, header in enumerate(headers_portfolio, 1):
            cell = ws2.cell(row=3, column=col_num, value=header)
            cell.font = s["header_font"]
            cell.fill = s["header_fill"]
            cell.alignment = s["align_center"]
            cell.border = s["border"]
        ws2.row_dimensions[3].height = 26

        for row_idx, p in enumerate(portfolios, 4):
            price_val = 0
            try:
                price_val = float(p.get('price', 0))
            except (ValueError, TypeError):
                pass
            area_val = p.get('area', '')
            try:
                area_val = int(area_val)
            except (ValueError, TypeError):
                pass

            row_data = [
                (p.get('title', ''), s["align_left"]),
                (price_val, s["align_right"]),
                (p.get('type', ''), s["align_center"]),
                (p.get('propertyType', ''), s["align_left"]),
                (p.get('rooms', ''), s["align_center"]),
                (area_val, s["align_center"]),
                (p.get('district', ''), s["align_left"]),
                (p.get('neighborhood', ''), s["align_left"]),
                (p.get('heating', ''), s["align_left"]),
                (p.get('age', ''), s["align_center"]),
                (p.get('status', ''), s["align_center"]),
                (p.get('sozlesme_tipi', ''), s["align_center"]),
                (p.get('sozlesme_bitis_tarihi', ''), s["align_center"]),
                (p.get('mulk_durumu', ''), s["align_center"]),
                (p.get('tapu_durumu_notlari', ''), s["align_left"]),
                (p.get('createdByName', ''), s["align_left"]),
            ]
            for col_num, (val, align) in enumerate(row_data, 1):
                cell = ws2.cell(row=row_idx, column=col_num, value=val)
                cell.font = s["data_font"]
                cell.alignment = align
                cell.border = s["border"]
                if col_num == 2:
                    cell.number_format = '₺#,##0'
            ws2.row_dimensions[row_idx].height = 20

        # Satıcı bölümü
        seller_start_row = len(portfolios) + 6
        ws2.cell(row=seller_start_row, column=1, value="SATICI MÜŞTERİ LİSTESİ").font = s["section_font"]
        ws2.row_dimensions[seller_start_row].height = 24

        header_seller_row = seller_start_row + 2
        headers_seller = [
            "Müşteri Adı", "Telefon", "E-posta", "Doğum Tarihi",
            "Sözleşme Tipi", "Sözleşme Bitiş Tarihi", "Mülk Durumu", "Tapu Notları",
            "Danışman", "Notlar"
        ]
        for col_num, header in enumerate(headers_seller, 1):
            cell = ws2.cell(row=header_seller_row, column=col_num, value=header)
            cell.font = s["header_font"]
            cell.fill = s["header_seller_fill"]
            cell.alignment = s["align_center"]
            cell.border = s["border"]
        ws2.row_dimensions[header_seller_row].height = 26

        for row_idx, sel in enumerate(sellers, header_seller_row + 1):
            row_data = [
                (sel.get('name', ''), s["align_left"]),
                (sel.get('phone', ''), s["align_center"]),
                (sel.get('email', ''), s["align_left"]),
                (sel.get('birthDate', ''), s["align_center"]),
                (sel.get('sozlesme_tipi', ''), s["align_center"]),
                (sel.get('sozlesme_bitis_tarihi', ''), s["align_center"]),
                (sel.get('mulk_durumu', ''), s["align_center"]),
                (sel.get('tapu_durumu_notlari', ''), s["align_left"]),
                (sel.get('createdByName', ''), s["align_left"]),
                (sel.get('notes', ''), s["align_left"]),
            ]
            for col_num, (val, align) in enumerate(row_data, 1):
                cell = ws2.cell(row=row_idx, column=col_num, value=val)
                cell.font = s["data_font"]
                cell.alignment = align
                cell.border = s["border"]
            ws2.row_dimensions[row_idx].height = 20

        _autofit_columns(ws2, skip_rows={1, 2, seller_start_row, seller_start_row + 1})

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)

        return send_file(
            out,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="crm_verileri_export.xlsx"
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        return {"error": str(e)}, 500
